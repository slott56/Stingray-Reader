"""
Workbook Implementations

See http://www.python-excel.org/ for a list of implementations.
Of these, we choose a few to handle XLSX, ODS, and Numbers.

"""
from collections.abc import Iterator
from pathlib import Path
from stingray.workbook import file_registry, Workbook, Sheet
from stingray.schema_instance import WBInstance, WBUnpacker
from typing import Optional, cast, Union, Any, IO, AnyStr
from types import TracebackType

try:
    import xlrd  # type: ignore [import]

    # We _could_ define an XLRD_Unpacker to work with the ``list[Cell]``
    # definitions for each row. This _could_ allow some more flexibility in
    # extracting the ``Cell.value`` and ``Cell.type`` information.
    # It, however, doesn't seem to be required.

    class XLSUnpacker(WBUnpacker):
        """Wrapper around :py:mod:`xlrd`."""
        def open(
            self,
            name: Path,
            file_object: Optional[Union[IO[str], IO[bytes]]] = None,
            **kwargs: Any,
        ) -> None:
            """
            Open an XLS file.

            :param name: :py:class:`Path` for the workbook
            :param file_object: Optional open file to use; if omitted, name is opened.
            :param kwargs: Additional KW Args to apply to the unpacker.
            """
            self.the_file = xlrd.open_workbook(name, **kwargs)

        def close(self) -> None:
            """Closes the XLS file."""
            if hasattr(self, "the_file") and self.the_file:
                del self.the_file

        def sheet_iter(self) -> Iterator[str]:
            """Yields names of the sheets in this workbook."""
            yield from (sheet.name for sheet in self.the_file.sheets())

        def instance_iter(self, name: str, **kwargs: Any) -> Iterator[WBInstance]:
            """
            Yields rows of this sheet of this workbook.

            :param name: Name of the sheet to process.
            :param kwargs: Additional keyword args (Ignored)
            :yields: :py:class:`WBInstance` rows from the workbook
            """
            xlrd_sheet = self.the_file.sheet_by_name(name)
            for row in xlrd_sheet.get_rows():
                yield cast(WBInstance, [cell.value for cell in row])

    @file_registry.file_suffix(".xls")
    class XLS_Workbook(Workbook[WBInstance]):
        """Facade for XLS Workbooks."""
        def __init__(self, name: Union[str, Path], **kwargs: Any) -> None:
            """
            Opens an XLS Workbook for access.

            :param name: :py:class:`Path` to workbook.
            :param kwargs: Additional args provided to :py:class:`XLSUnpacker`.
            """
            super().__init__(name)
            self.xlrd_args = kwargs
            self.unpacker = XLSUnpacker()
            self.unpacker.open(self.name, None, **self.xlrd_args)

        def close(self) -> None:
            """Closes this workbook."""
            self.unpacker.close()


except ImportError:  # pragma: no cover
    pass


try:
    from openpyxl import load_workbook  # type: ignore [import]
    import openpyxl.cell.cell  # type: ignore [import]

    class XLSXUnpacker(WBUnpacker):
        """Wrapper around :py:mod:`openpyxl`."""
        def open(
            self,
            name: Path,
            file_object: Optional[Union[IO[str], IO[bytes]]] = None,
            **kwargs: Any,
        ) -> None:
            """
            Opens the XLSX file.

            :param name: :py:class:`Path` to the workbook.
            :param file_object: Optional open file to use; if omitted, name is opened.
            :param kwargs: KW Args provided to ``openpyxl``
            """
            self.the_file = load_workbook(filename=name, **kwargs)

        def close(self) -> None:
            """Closes the XLSX file."""
            if hasattr(self, "the_file") and self.the_file:
                self.the_file.close()
                del self.the_file

        def sheet_iter(self) -> Iterator[str]:
            """yields sheet names from this workbook."""
            return (cast(str, name) for name in self.the_file.sheetnames)

        def instance_iter(self, name: str, **kwargs: Any) -> Iterator[WBInstance]:
            """
            Yields rows from the given sheet.

            :param name: sheet name
            :param kwargs: additional kw args (not used)
            :yields: :py:class:`WBInstance` rows from the workbook
            """
            pyxl_sheet = self.the_file[name]
            for row in pyxl_sheet.iter_rows():
                yield cast(WBInstance, [cell.value for cell in row])

    @file_registry.file_suffix(".xlsx")
    class XLSX_Workbook(Workbook[WBInstance]):
        """Facade for XLSX Workbooks."""
        def __init__(self, name: Union[str, Path], **kwargs: Any) -> None:
            """
            Opens an XLSX workbook for access.

            :param name: :py:class:`Path` to the workbook
            :param kwargs: Additional kw args provided to the unpacker.
            """
            super().__init__(name)
            self.kwargs = kwargs
            self.unpacker = XLSXUnpacker()
            self.unpacker.open(self.name, None, **self.kwargs)

        def close(self) -> None:
            """Closes the workbook."""
            self.unpacker.close()


except ImportError:  # pragma: no cover
    pass


try:
    import pyexcel_ods3  # type: ignore [import]

    class ODSUnpacker(WBUnpacker):
        """Wrapper around :py:mod:`pyexcel_ods3`."""
        def open(
            self,
            name: Path,
            file_object: Optional[Union[IO[str], IO[bytes]]] = None,
            **kwargs: Any,
        ) -> None:
            """
            Opens the ODS file.

            :param name: :py:class:`Path` to the workbook.
            :param file_object: Optional open file to use; if omitted, name is opened.
            :param kwargs: KW Args provided to ``openpyxl``
            """
            self.the_file = pyexcel_ods3.get_data(str(name), **kwargs)

        def close(self) -> None:
            """Closes the ODS file."""
            if hasattr(self, "the_file") and self.the_file:
                del self.the_file

        def sheet_iter(self) -> Iterator[str]:
            """Yields names of the sheets in this workbook."""
            return iter(self.the_file.keys())

        def instance_iter(self, name: str, **kwargs: Any) -> Iterator[WBInstance]:
            """
            Yields individual rows from the current sheet.

            :param name: Sheet name
            :param kwargs: Additional kw args (not used)
            :yields: :py:class:`WBInstance` of worksheet rows.
            """
            pyexcel_sheet = self.the_file[name]
            for row in pyexcel_sheet:
                yield cast(WBInstance, row)

    @file_registry.file_suffix(".ods")
    class ODS_Workbook(Workbook[WBInstance]):
        """Facade for XLSX Workbooks."""
        def __init__(self, name: Union[str, Path], **kwargs: Any) -> None:
            """
            Opens the workbook for access.

            :param name: :py:class:`Path` to the workbook
            :param kwargs: Additional keyword args for the Unpacker
            """
            super().__init__(name)
            self.kwargs = kwargs
            self.unpacker = ODSUnpacker()
            self.unpacker.open(self.name, None, **self.kwargs)

        def close(self) -> None:
            """Closes the workbook."""
            self.unpacker.close()


except ImportError:  # pragma: no cover
    pass


try:
    import numbers_parser  # type: ignore [import]

    class NumbersUnpacker(WBUnpacker):
        """Wrapper around :py:mod:`numbers_parser`."""
        def open(
            self,
            name: Path,
            file_object: Optional[Union[IO[str], IO[bytes]]] = None,
            **kwargs: Any,
        ) -> None:
            """
            Opens the numbers file.

            :param name: :py:class:`Path` to the workbook.
            :param file_object: Optional open file to use; if omitted, name is opened.
            :param kwargs: KW Args provided to ``numbers_parser``
            """
            self.the_file = numbers_parser.Document(name, **kwargs)

        def close(self) -> None:
            """Closes the numbers file."""
            if hasattr(self, "the_file") and self.the_file:
                del self.the_file

        def sheet_iter(self) -> Iterator[str]:
            """Yields the names of sheets and tables.
            The names take the form of sheet::table.
            """
            return (
                f"{sheet.name}::{table.name}"
                for sheet in self.the_file.sheets()
                for table in sheet.tables()
            )

        def instance_iter(self, name: str, **kwargs: Any) -> Iterator[WBInstance]:
            """
            Yields individual rows from the current sheet.

            :param name: Sheet::Table composite name
            :param kwargs: Additional kw args (not used)
            :yields: :py:class:`WBInstance` of worksheet rows.
            """
            sheet, _, table = name.partition("::")
            numbers_sheet = self.the_file.sheets()[sheet].tables()[table]
            for row in numbers_sheet.iter_rows():
                yield cast(WBInstance, [cell.value for cell in row])

    @file_registry.file_suffix(".ods")
    class Numbers_Workbook(Workbook[WBInstance]):
        """Facade for XLSX Workbooks."""
        def __init__(self, name: Union[str, Path], **kwargs: Any) -> None:
            """
            Opens the workbook for access.

            :param name: :py:class:`Path` to the workbook
            :param kwargs: Additional keyword args for the Unpacker
            """
            super().__init__(name)
            self.numbers_args = kwargs
            self.unpacker = NumbersUnpacker()
            self.unpacker.open(self.name, None, **self.kwargs)

        def close(self) -> None:
            """Closes the workbook."""
            self.unpacker.close()


except ImportError:  # pragma: no cover
    pass
